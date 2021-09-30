import openpyxl


def read_file(file_path):
    wookbook = openpyxl.load_workbook(file_path)
    worksheet = wookbook.active
    dictionary = {}
    table_name = []
    columns = []
    insertion = []

    for i in range(0, worksheet.max_row):
        data = []

        for col in worksheet.iter_cols(1, worksheet.max_column):
            if i == 0 and col[i].value != None:
                table_name = col[i].value
            elif i == 1:
                columns.append(col[i].value)
            else:
                data.append(col[i].value)
        
        if data and data[0] != None:
            insertion.append(data)
            # zip_iterator  = zip(columns, data)
            # insertion.append(dict(zip_iterator))
            
    dictionary["TableName"] = table_name
    dictionary["Columns"] = columns
    dictionary["Insertion"] = insertion

    return dictionary


def create_query(file_path):
    result = read_file(file_path)
    
    columns_str = "("
    for i in result["Columns"]:
        columns_str += "'" + i + "'" + ','
    str_last_index = len(columns_str) - 1
    columns_str = columns_str[:str_last_index] + ")"

    values_str_list = []
    for j in result["Insertion"]:
        values_str = "("
        for k in j:
            values_str += "'" + str(k) + "'" + ','
        str_last_index = len(values_str) - 1
        values_str = values_str[:str_last_index] + ")"
        values_str_list.append(values_str)

    f = open("output/query.txt","w+")
    for l in range(len(values_str_list)):
        f.write(f'INSERT INTO {result["TableName"]} {columns_str} VALUES {values_str_list[l]};\n')


def main():
    create_query("input/input.xlsx")


if __name__ == "__main__":
    main()